#! /usr/bin/python
### To compare Qualys scans from the beginning and middle of the month
### and report the new PCI vulnerabilities discovered, identifying any expired false positives

import os
import argparse
import openpyxl
from openpyxl import load_workbook

parser = argparse.ArgumentParser(prog='comarePCIscans')
parser.add_argument('-D', '--Debug', help='Debug Mode assists in determining issues being raised by the script.', action='store_true')
parser.add_argument('-F', '--firstScan', help='first scan (original scan)')
parser.add_argument('-N', '--newScan', help='new scan')
parser.add_argument('-fP', '--falsePositives', help='false positives file')
parser.add_argument('-V', '--vulnerabilitiesFile', help='path and name of new vulnerabilities file')
args = parser.parse_args()


# first scan (base)
pciA = load_workbook(args.firstScan, data_only=True)
scanA = pciA.get_sheet_by_name(pciA.get_sheet_names()[0])  #sheet
maxrA = str(scanA.max_row - 2)

# second scan (identify new vulns in this scan)
pciB = load_workbook(args.newScan, data_only=True)
scanB = pciB.get_sheet_by_name(pciB.get_sheet_names()[0]) #sheet
maxrB = str(scanB.max_row - 2)

# false positives
falseP = load_workbook(args.falsePositives, data_only=True)
falseS = falseP.get_sheet_by_name(falseP.get_sheet_names()[0]) #sheet
maxrF = str(falseS.max_row)


def vulnDict(scan, maxrow):
	dscan = {}
	for ipRow in scan['A8': 'A' + str(maxrow)]:
		for ipCell in ipRow:
			if str(scan['AA' + str(ipCell.row)].value) == 'yes':
				if str(scan['H' + str(ipCell.row)].value) != 'Ig':
					if str(ipCell.value) +':'+ str(scan['J' +str(ipCell.row)].value) in dscan:
						dscan[(str(ipCell.value) +':'+ str(scan['J' +str(ipCell.row)].value))].append(scan['G' +str(ipCell.row)].value)
					else:
						dscan[str(ipCell.value) +':'+ str(scan['J' +str(ipCell.row)].value)] = [scan['G' +str(ipCell.row)].value]

	if args.Debug:
		print('[+] Scan Analyzed')

	return dscan


### MAIN
dscanA = vulnDict(scanA, maxrA)
dscanB = vulnDict(scanB,maxrB)

### WriteReport
# new vulnerabilities report
vulnP = args.vulnerabilitiesFile
vulnF = openpyxl.Workbook()
newVulns = vulnF.get_active_sheet()

#add column titles
newVulns['A1'].value = 'IP:Port'
newVulns['B1'].value = 'Vuln Title'
newVulns['C1'].value = 'Port'
newVulns['D1'].value = 'Status'
if args.Debug:
	print('[+] Workbook Created and Column Headers Complete')

# compare dictionaries and write the newly discovered vulns
i = 2
for key in dscanB:
	if key in dscanA:
		for value in dscanB[key]:
			if value in dscanA[key]:
				pass
			else:
				newVulns['A' + str(i)].value = key.split(':')[0]
				newVulns['B' + str(i)].value = value
				newVulns['C' + str(i)].value = key.split(':')[1]
				i += 1
	else:
		for value in dscanB[key]:
			newVulns['A' + str(i)].value = key.split(':')[0]
			newVulns['B' + str(i)].value = value
			newVulns['C' + str(i)].value = key.split(':')[1]
			i += 1
if args.Debug:
	print('[+] Discovered Vulnerabilities writen to file')

# compare new vulns to False Positives and write the label
dfalseS = {}
dnewVulns = {}
maxrV = str(newVulns.max_row)

for ipRow in newVulns['A2': 'A' + maxrV]:
	for ipCell in ipRow:
		ip = str(ipCell.value)
		if ip in dnewVulns:
			dnewVulns[ip].append(newVulns['B' + str(ipCell.row)].value)
		else:
			dnewVulns[ip] = [newVulns['B' + str(ipCell.row)].value]

for iprowF in falseS['C2': 'C' + maxrF]:
	for ipcellF in iprowF:
		if str(falseS['F' + str(ipcellF.row)].value) != 'Rejected':
			if str(ipcellF.value) in dnewVulns:
				if str(ipcellF.value) in dfalseS:
					dfalseS[str(ipcellF.value)].append(falseS['B' + str(ipcellF.row)].value)
				else:
					dfalseS[str(ipcellF.value)] = [falseS['B' + str(ipcellF.row)].value]

for key in dnewVulns:
	for value in dnewVulns[key]:
		if key in dfalseS:
			if value in dfalseS[key]:
				# print(key, value)
				for ipRow in newVulns['A2': 'A' + maxrV]:
					for ipCell in ipRow:
						if str(ipCell.value) == key:
							if value in newVulns['B' + str(ipCell.row)].value:
								newVulns['D' + str(ipCell.row)].value = 'FP: Check Port'


if args.Debug:
	print('[+] False Positives labeled')

vulnF.save(vulnP)
if args.Debug:
	print('[+] New Report Saved')

